from datetime import datetime
from datetime import timedelta
import pytz
from .models import Station, Schedule, get_version, AggregationUnit, Company, CommerceReport, get_commerce_version
import openpyxl
import xlsxwriter
import xml.etree.ElementTree as ET


def day_duration(day: datetime) -> int:
    """
    Returns the duration of a day in hours (24, 23 or 25 hours).
    """
    d = day.replace(hour=0, minute=0, second=0, microsecond=0)
    # convert kyiv to UTC
    start_utc = pytz.timezone('Europe/Kyiv').localize(d).astimezone(pytz.timezone('UTC'))
    end_utc = (pytz.timezone('Europe/Kyiv').localize(d + timedelta(days=1))).astimezone(pytz.timezone('UTC'))
    if start_utc.hour == end_utc.hour:
        return 24
    elif start_utc.hour < end_utc.hour:
        return 25
    else:
        return 23
    


def upload_schedule(file, user):
    """
    Uploads schedule from file.
    :param file: file with schedule
    :return: result of upload (success or error message)
    """
    # read excel file
    if not file.name.endswith('.xlsx'):
        return 'File must be an Excel (.xlsx) file.'

    wb = openpyxl.load_workbook(file)
    ws = wb['data']
    # check if have authority to upload for station (and station exists)
    # check if not over deadline
    for row in ws.iter_rows(min_row=2, values_only=True):
        # TODO: mms name to lowercase to avoid case issues
        station = Station.objects.filter(mms_name=row[0]).first()
        if not station:
            return f'Station {row[0]} not found'
        # str to datetime
        date = row[1]
        duration = day_duration(date)
        if duration + 2 > len(row):  # 2 for mms_name and date 
            print(f'Invalid schedule for {station.alias} on {date}. Expected {duration} hours, got {len(row) - 2}.')
            return f'Invalid schedule for {station.alias} on {date}. Expected {duration} hours, got {len(row) - 2}.'
        version = get_version(station, date)
        # read schedule datax
        schedule_data = {}
        for i in range(duration):
            val = row[i + 2]
            if val is None:
                print(f'Invalid schedule for {station.alias} on {date}. Value at hour {i} is None.')
                return f'Invalid schedule for {station.alias} on {date}. Value at hour {i} is None.'
            schedule_data[i] = val
            # TODO: check if value is valid
        schedule = Schedule(
            station=station,
            date=date,
            version=version,
            schedule=schedule_data,
            created_by=user
        )
        schedule.save()
        print(f'Uploaded schedule for {station.alias} on {date} (v{version})')
    return 'Success'    
        # TODO: implement validation checks
        # check if values count is correct for date
        # check if values is correct
        # check if no double stations in file
        # upload schedule in bulk
        # upload only if data change
        # upload only for future dates
        # error if same schedule (company and date not unique)

# TODO: generate report
def generate_report(date: datetime, user):
    # report is per company that owns the station and separately by station
    # get only latest version for each date, station pair
    # find unique stations for date, than find latest schedule for each station
    schedules = Schedule.objects.filter(date=date)
    unique_stations = schedules.values('station').distinct()
    print(f'Unique stations for {date}: {unique_stations}')
    # init dict to store report data
    #TODO: add total row for stations and companies
    report_data = {'companies': {}, 'stations': {}, 'total': {}}
    for station in unique_stations:
        schedule = Schedule.objects.filter(station__pk=station['station'], date=date).order_by('-version').first()
        # add station schedule to report data
        report_data['stations'][str(schedule.station.pk)] = schedule.schedule
        # add company schedule to report data
        if report_data['companies'].get(str(schedule.station.company.pk)) is None:
            report_data['companies'][str(schedule.station.company.pk)] = schedule.schedule.copy()
        else:
            for hour, value in schedule.schedule.items():
                report_data['companies'][str(schedule.station.company.pk)][hour] += value
        # add to total row schedule
        for hour, value in schedule.schedule.items():
            if report_data['total'].get(hour) is None:
                report_data['total'][hour] = value
            else:
                report_data['total'][hour] += value

    # TODO: only if data changed
    report = CommerceReport.objects.filter(date=date).order_by('-version').first()
    
    report = create_report_if_changed(date, report_data, report, user)
    save_commerce_report(report)

    print('Report data:')
    for company, schedule in report_data['companies'].items():
        print(f'Company: {company}, Schedule: {schedule}')
    for station, schedule in report_data['stations'].items():
        print(f'Station: {station}, Schedule: {schedule}')

def create_report_if_changed(date, report_data, report, user):
    if report and report.report == report_data:
        print(f'No changes detected for {date}. Skipping report saving')
    else:
        version = get_commerce_version(date)
        report = CommerceReport(
            date=date,
            version=version,
            report=report_data,
            created_by=user
        )
        report.save()
    return report


def save_commerce_report(report: CommerceReport):
    # save excel file with report data
    wb = xlsxwriter.Workbook(f'/Users/tritaer/programing/mergo/commerce_report_{report.date.strftime("%Y-%m-%d")}_v{report.version}.xlsx')
    ws_company = wb.add_worksheet('Companies')
    ws_station = wb.add_worksheet('Stations')

    bold = wb.add_format({'bold': True})
    date_format = wb.add_format({'num_format': 'yyyy-mm-dd'})
    # formatting ws_company
    ws_company.set_row(0, None, bold)
    ws_company.set_column(0, 0, 10, date_format)  # Date
    ws_company.set_column(1, 1, 18)  # X
    ws_company.set_column(2, 2, 20)  # mmsName
    # ws_company
    # header
    ws_company.write(0, 0, 'Дата')
    ws_company.write(0, 1, 'X')
    ws_company.write(0, 2, 'mmsName')
    for i in range(25):
        ws_company.write(0, 3 + i, i)
    # data
    for row, (company_pk, schedule) in enumerate(report.report['companies'].items(), start=1):
        ws_company.write(row, 0, report.date)
        company = Company.objects.get(pk=company_pk)
        ws_company.write(row, 1, company.x_code)
        ws_company.write(row, 2, company.mms_name)
        for hour, value in schedule.items():
            ws_company.write(row, 3 + int(hour), value / 1000)
    # total row
    total_row = len(report.report['companies']) + 1
    ws_company.write(total_row, 2, 'Загально', bold)
    for hour, value in report.report['total'].items():
        ws_company.write(total_row, 3 + int(hour), value / 1000, bold)
    # formatting ws_station
    ws_station.set_row(0, None, bold)
    ws_station.set_column(0, 0, 10, date_format)  # Date
    ws_station.set_column(1, 1, 18)  # W
    ws_station.set_column(2, 2, 20)  # mmsName
    # ws_station
    # header
    ws_station.write(0, 0, 'Дата')
    ws_station.write(0, 1, 'W')
    ws_station.write(0, 2, 'mmsName')
    for i in range(25):
        ws_station.write(0, 3 + i, i)
    # data
    for row, (station_pk, schedule) in enumerate(report.report['stations'].items(), start=1):
        ws_station.write(row, 0, report.date)
        station = Station.objects.get(pk=station_pk)
        ws_station.write(row, 1, station.w_code)
        ws_station.write(row, 2, station.mms_name)
        for hour, value in schedule.items():
            ws_station.write(row, 3 + int(hour), value / 1000)
    # total row
    total_row = len(report.report['stations']) + 1
    ws_station.write(total_row, 2, 'Загально', bold)
    for hour, value in report.report['total'].items():
        ws_station.write(total_row, 3 + int(hour), value / 1000, bold)
    wb.close()


